import openpyxl  # работать с эксель
import re  # искать паттерны
from datetime import datetime

import time

# Устанавливаем таймаут на 5 секунд
timeout = 5

# C:\Users\Admin\PycharmProjects\pythonProject_wb\openpyxl_wb.py

'''
    (!) что важно исправить в первую очередь:
        + грамотно записывать пустые значения (None) из строки, чтобы им был выделен отдельный столбец
        по факту: выделяется
        
    (!) Неопознанный товар - обрабатывать с предупреждением (сейчас просто записывается)
    
    (!) убирать перенос строки у 
            "Печенье
        "
        странное
'''

''' как можно улучшать:
    + пустые строки искать не через первый элемент, 
        а через первый НЕнулевой элемент, 
        тогда можно выцепить и строки, 
        где значимые данные не в первом столбюце
    + наименование столбцов искать паттерном, но последовательность определять фактически,
        т.е. подстраховаться от перестановки столбцов
    + (!) выводить в лог warning, если поменялся формат файла, поставить проверки        
'''

# обработка нескольких файлов
print('Программа для обработки актов приемки WB.')
print(' Далее укажите расположение папки с актами (соберите все акты в одну папку).')
print('  Пример расположения: C:\\Users\\Admin\\Downloads\\Акты приемки')
print(' Пример 2: C:\\Users\\Admin\\Downloads\\Акт приемки 777\\')
# для копипаста: C:\Users\Admin\Downloads\Акт приемки 777\act-income-777.xlsx
# print(' Пример: C:\\Users\\Admin\\Downloads\\Акт приемки 777\\act-income-777.xlsx')
print('  Примечание: можно оставлять обратный слеш в конце пути к папке, это ок\\')
print(' ')
directory_location = input("Введите расположение папки с актами на вашем ПК: ")
directory_location_edit = directory_location.replace("\\", "\\\\")  # экранируем слеши

from pathlib import Path

files = [str(file) for file in Path(directory_location_edit).iterdir() if file.is_file()]

'''
записать первую строку с названием столбцов:

table = []
table_row = []
for j in range(8):
    table_row.append(row[j+1])

print('  (ок) Найдена строка с товаром. Номер пп: ', item_order, ' в строке ', idx)
                    print('Товар (наименование)', row[1])
                    print('Ед. изм.', row[2])
                    print('баркод', row[3])
                    print('артикул продавца', row[4])
                    print('сорт, размер', row[5])
                    print('КИЗ', row[6])
                    if isinstance(row[6], str) and row[6] == '':
                        print(' (i) Заполним КИЗ нулем: 0')
                        row[6] = 0
                    print('ШК короба', row[7])
                    print('кол-во', row[8])
                    
                    '''

file_count = 0
non_file_count = 0

table2 = []

for file in files:
    pattern_type = re.compile(r"\.(xlsx|xlsm|xltx|xltm)$")
    match_type = re.search(pattern_type, file)

    if match_type:
        print('Обрабатывается файл: ', file)
        file_count +=1
        # Открываем файл
        workbook = openpyxl.load_workbook(file)  # file_location_edit)

        # Получаем активный лист
        sheet = workbook.active

        # паттерны для поиска
        pattern_act_wb = r"Акт приемки товара № (\d+)"  # для поиска номера акта
        pattern_item_wb = r'(\d+)' # для поиска строк с поставками
        # риск: можно попасть на строку с номерами столбцов: (None, '1', '2', '3', '4', '5', '6', '7', '8')
        # но пока алгоритм берет первый элемент - так что - ок, такая строка нам не подойдет
        pattern_date_wb = r'"(\d+)"\s+(\d+)\s+(\d+)г\.'  # "22" 12 2023г.

        # переменные для наполнения
        act_number = None
        act_date = None
        act_date_format = None
        table = []



        # print('Мы внутри sheet.iter_rows, смотрим кортежи строк, строки экселя')
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            print(f" Строка {idx} содержит следующее: {row}")
            # сейчас ищем только в 1м элементе кортежа строк
            row_string = row[0]
            print('  Первый элемент строки: ', row_string)

            if isinstance(row_string, str) and act_number is None:  # если первый элемент - строка (т.е. кортеж не пуст)
                print('   Т.к. первый элемент - строка, и мы ещё не нашли номер акта, то мы внутри isinstance, ищем совпадение с паттерном НОМЕР АКТА')
                found_pattern_act = re.search(pattern_act_wb, row_string)
                if found_pattern_act:
                    act_number = found_pattern_act.group(1)  # у нас один паттерн, достаем найденное им, если есть
                    print('  (ок) Найден номер акта: ', act_number, ' в строке ', idx)
                else:
                    print(' не нашли тут номера акта.')  # после добавление act_number is None - сообщения не будет такого

            if isinstance(row_string, int):  # если первый элемент - число (т.е. кортеж не пуст)
                print('   Т.к. первый элемент - число, то мы внутри isinstance, ищем совпадение с паттерном ЗАПИСЬ ТОВАРА')
                found_pattern_item = re.search(pattern_item_wb, str(row_string))
                if found_pattern_item:
                    item_order = found_pattern_item.group(1)  # у нас один паттерн, достаем найденное им, если есть
                    print('  (ок) Найдена строка с товаром. Номер пп: ', item_order, ' в строке ', idx)
                    print('')
                    print('Товар (наименование)', row[1])
                    print('Ед. изм.', row[2])
                    print('баркод', row[3])
                    print('артикул продавца', row[4])
                    print('сорт, размер', row[5])
                    print('КИЗ', row[6])

                    print('ШК короба', row[7])
                    print('кол-во', row[8])

                    table_row = []
                    table_row.append(act_number)
                    for j in range(5):
                        table_row.append(row[j+1])

                    print('')
                    kiz = row[6]
                    if isinstance(row[6], str) and row[6] == '':
                        print(' (i) Заполним КИЗ нулем: 0')
                        kiz = 0

                    for j in range(6,8):
                        table_row.append(row[j+1])

                    table.append(table_row)
                    print('')
                    print('Записали в таблицу и получили итого: ', table)
                    print('')

                else:
                    print(' не нашли в этой строке товара.')

            if isinstance(row_string, type(None)):
                for col_index, col_value in enumerate(row, start=1):
                    if not isinstance(col_value, type(None)):
                        found_pattern_date = re.search(pattern_date_wb, col_value)
                        if found_pattern_date:
                            # Преобразование в дату
                            date_obj = datetime.strptime(col_value, '"%d" %m %Yг.')

                            # Форматирование в требуемый формат
                            formatted_date = date_obj.strftime('%d.%m.%Y')
                            act_date = col_value
                            act_date_format = formatted_date
                            print('Нашли дату акта: ',col_value)
                            print('Преобразовали в дату вида: ', act_date_format)

            if isinstance(row_string, str) and row_string=='Итого':
                amount_total_row=row[8]
                print('Нашли итого:', amount_total_row)

                counter_total_goods = 0
                for table_row_2 in table:
                    counter_total_goods += table_row_2[7]

                if amount_total_row == counter_total_goods:
                    print('В С ё  В  П О Р Я Д К Е : итого совпал с суммой товаров.')

                else:

                    print(' П Р О Б Л Е М А : итого НЕ совпал с суммой товаров!')
                    # Запоминаем время начала
                    start_time = time.time()

                    flag_cycle = 1
                    # Выполняем какую-то работу
                    while flag_cycle == 1:
                        # Проверяем, прошло ли 5 секунд
                        if time.time() - start_time >= timeout:
                            print("Прошло 5 секунд")
                            flag_cycle = 0
                        else:
                            print("Работаем...")

        table_row_count=len(table)
        print('')
        print('Идем дальше. Добавляем дату в таблицу')
        for row_data in (table):
            row_data.append(act_date_format)
            print(table)

        print('')
        print('Обработали файл.')
        print('')

        table2 = table2+table
        print('присоединяем таблицу')
        print('итоговая на данный момент: ', table2)

    else:
        print('')
        print('Пропускаем неподходящий файл: ', file)
        non_file_count +=1

print('')
print('Файлов больше нет - завершаем работу.')
print('Итого обработано файлов эксель: ', file_count)
print('Итого пропущено файлов из-за формата: ', non_file_count)
print('Спасибо')

print(' ')
print('p.s. и сохраним таблицу в файл эксель')

from openpyxl import Workbook

# Создаем новую книгу Excel
workbook = Workbook()

# Получаем активный лист
sheet = workbook.active

# Записываем данные в лист
for row in table2:
    sheet.append(row)

# Сохраняем книгу Excel в файл
workbook.save(filename="wb_act_handmade_20240302.xlsx")

print('сохранили, завершаем.')

'''
# Переменные для количества строк и столбцов
num_rows = 5  # Например, 5 строк
num_cols = 11  # 11 столбцов

# Создаем пустую таблицу
table = []

# Добавляем строки в таблицу
for i in range(num_rows):
    row = []  # Создаем пустую строку
    for j in range(num_cols):
        row.append(0)  # Добавляем в строку пустые значения (в этом примере 0)
    table.append(row)  # Добавляем строку в таблицу

# Выводим таблицу
for row in table:
    print(row)
'''


''' попробовать Пандас:

 import pandas as pd

# Создание DataFrame (пример)
data = {'A': [1, 2, 3], 'B': [4, 5, 6]}
df = pd.DataFrame(data)

num_rows = df.shape[0]
print("Число строк в таблице:", num_rows)

 '''